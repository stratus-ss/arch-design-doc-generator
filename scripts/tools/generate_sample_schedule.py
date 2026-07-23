#!/usr/bin/env python3
"""
Generate a synthetic Migration_Weekly_Schedule.xlsx with fictional data.

Produces a realistic-looking migration schedule workbook suitable for
inclusion in the published repository as sample/demo data. All VM names,
DNS names, datacenter names, cluster names, and annotations are fictional.

Usage:
    python generate_sample_schedule.py -o Sample_Migration_Weekly_Schedule.xlsx
"""

import argparse
import hashlib
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "Triage",
    "Migration_Status", "Backup_Completed", "JIRA", "Point_of_Contact",
    "Plan", "Wave_ID", "Migration_Type", "Design_Review", "Post_Migration_Validation",
    "VM_Name", "DNS_Name", "Powerstate", "CPUs", "Memory_GB",
    "Disk_Capacity_GB", "Datacenter", "Cluster", "Resource_Pool",
    "OS", "Latency_Sensitivity", "CBT", "VM_ID", "Annotation",
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

TRIAGE_FILLS = {
    "migrate": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "rebuild": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "retire":  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
TRIAGE_WEIGHTS = [75, 15, 10]

SITES = {
    "Site Alpha": {"dc": "ALPHA", "prefix": "SA", "vm_count": 80},
    "Site Alpha Lab": {"dc": "ALPHA_LAB", "prefix": "SAL", "vm_count": 30},
    "Site Alpha Tier 0": {"dc": "ALPHA_ISE", "prefix": "SAT", "vm_count": 25},
    "Site Beta": {"dc": "BETA", "prefix": "SB", "vm_count": 90},
    "Site Beta Tier 0": {"dc": "BETA_ISE", "prefix": "SBT", "vm_count": 30},
    "Site Gamma Lab": {"dc": "GAMMA_LAB", "prefix": "SGL", "vm_count": 70},
    "Site Gamma Lab Tier 0": {"dc": "GAMMA_LAB_ISE", "prefix": "SGT", "vm_count": 25},
    "Site Delta": {"dc": "DELTA", "prefix": "SD", "vm_count": 40},
}

OS_LIST = [
    "Microsoft Windows Server 2019 (64-bit)",
    "Microsoft Windows Server 2022 (64-bit)",
    "Microsoft Windows 10 (64-bit)",
    "Microsoft Windows 11 (64-bit)",
    "Red Hat Enterprise Linux 8 (64-bit)",
    "Red Hat Enterprise Linux 9 (64-bit)",
    "Ubuntu Linux (64-bit)",
    "SUSE Linux Enterprise 15 (64-bit)",
    "Other 3.x or later Linux (64-bit)",
    "Debian GNU/Linux 12 (64-bit)",
    "CentOS Stream 9 (64-bit)",
    "Oracle Linux 8 (64-bit)",
]

ROLES = [
    "web", "app", "db", "cache", "proxy", "queue", "etl", "api",
    "auth", "log", "mon", "dns", "ntp", "smtp", "ftp", "vpn",
    "ldap", "sso", "ci", "cd", "repo", "scan", "vault", "jump",
]

ENVS = ["prod", "dev", "stg", "qa", "uat", "dr"]

CLUSTER_SUFFIXES = [
    "PROD_GENERAL", "PROD_HIGH_SECURITY", "PROD_APPLIANCES",
    "DEV_GENERAL", "INFRASTRUCTURE", "WORKSTATIONS",
]

ANNOTATIONS = [
    None,
    "Application server for internal portal",
    "Database replica — read-only workloads",
    "Monitoring agent collector node",
    "CI/CD runner for build pipelines",
    "Load balancer health-check endpoint",
    "Log aggregation and SIEM forwarding",
    "Backup proxy for nightly snapshot jobs",
    "Identity management service",
    "Certificate management appliance",
    "Container image registry mirror",
    "Network packet broker",
    "Automated testing harness",
    "Configuration management server",
    "Service mesh control plane",
]


def _deterministic_seed(site_name: str, index: int) -> int:
    h = hashlib.md5(f"{site_name}-{index}".encode()).hexdigest()
    return int(h[:8], 16)


def generate_vm(site_name: str, site_cfg: dict, index: int) -> list:
    rng = random.Random(_deterministic_seed(site_name, index))

    role = rng.choice(ROLES)
    env = rng.choice(ENVS)
    seq = f"{index:02d}"
    prefix = site_cfg["prefix"].lower()
    vm_name = f"{prefix}-{role}-{env}-{seq}"
    dns_name = f"{vm_name}.example.corp" if rng.random() > 0.15 else None

    powerstate = "poweredOn" if rng.random() > 0.12 else "poweredOff"
    cpus = rng.choice([2, 4, 8, 16, 24, 32])
    mem_gb = rng.choice([4, 8, 16, 32, 64, 128, 256])
    disk_gb = round(rng.choice([50, 100, 200, 500, 1000, 2000]) * rng.uniform(0.8, 1.2), 2)

    dc = site_cfg["dc"]
    cluster_suffix = rng.choice(CLUSTER_SUFFIXES)
    cluster = f"{dc}_{cluster_suffix}"
    resource_pool = f"/{dc}/{cluster}/Resources"

    os_name = rng.choice(OS_LIST)
    latency = rng.choice(["normal", "normal", "normal", "low"])
    cbt = rng.choice([None, None, "TRUE"])
    vm_id = f"vm-{rng.randint(100, 9999999)}"
    annotation = rng.choice(ANNOTATIONS)

    triage = rng.choices(["migrate", "rebuild", "retire"], weights=TRIAGE_WEIGHTS)[0]
    migration_type = rng.choices(["warm", "cold"], weights=[90, 10])[0] if triage == "migrate" else ""

    return [
        triage,
        "", "", "", "",
        "", "", migration_type, "", "",
        vm_name, dns_name, powerstate, cpus, mem_gb, disk_gb,
        dc, cluster, resource_pool, os_name, latency, cbt,
        vm_id, annotation,
    ]


def main():
    parser = argparse.ArgumentParser(
        description="Generate a synthetic sample Migration Weekly Schedule XLSX"
    )
    parser.add_argument(
        "-o", "--output",
        default="Sample_Migration_Weekly_Schedule.xlsx",
        help="Output XLSX path",
    )
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    for site_name, site_cfg in SITES.items():
        ws = wb.create_sheet(title=site_name[:31])

        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        triage_col = HEADERS.index("Triage") + 1

        for i in range(site_cfg["vm_count"]):
            row_data = generate_vm(site_name, site_cfg, i)
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=i + 2, column=col_idx, value=val)
            triage_val = row_data[triage_col - 1]
            if triage_val in TRIAGE_FILLS:
                ws.cell(row=i + 2, column=triage_col).fill = TRIAGE_FILLS[triage_val]

        for col_idx in range(1, len(HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(HEADERS[col_idx - 1]) + 4, 12)
        ws.freeze_panes = "A2"

    wb.save(args.output)
    total = sum(s["vm_count"] for s in SITES.values())
    print(f"Generated {args.output}: {len(SITES)} sheets, {total} sample VMs")


if __name__ == "__main__":
    main()
