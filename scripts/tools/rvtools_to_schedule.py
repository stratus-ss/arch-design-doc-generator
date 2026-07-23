#!/usr/bin/env python3
"""
RVTools to Migration Schedule Generator

Reads RVTools vInfo exports and produces a migration weekly schedule XLSX
with per-site sheets, matching the prior engagement source-of-truth format adapted
for {CLIENT}.

Usage:
    python rvtools_to_schedule.py ./RVTools/*.xlsx -o Migration_Weekly_Schedule.xlsx
    python rvtools_to_schedule.py file1.xlsx file2.xlsx --exclude-powered-off
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


TRACKING_COLUMNS = [
    "Migration_Status",
    "Backup_Completed",
    "JIRA",
    "Point_of_Contact",
    "Plan",
    "Wave_ID",
    "Design_Review",
    "Post_Migration_Validation",
]

RVTOOLS_TO_OUTPUT = [
    ("VM", "VM_Name"),
    ("DNS Name", "DNS_Name"),
    ("Powerstate", "Powerstate"),
    ("CPUs", "CPUs"),
    (None, "Memory_GB"),  # computed from Memory (MB)
    (None, "Disk_Capacity_GB"),  # computed from Total disk capacity MiB
    ("Datacenter", "Datacenter"),
    ("Cluster", "Cluster"),
    ("Resource pool", "Resource_Pool"),
    ("OS according to the configuration file", "OS"),
    ("Latency Sensitivity", "Latency_Sensitivity"),
    ("CBT", "CBT"),
    ("VM ID", "VM_ID"),
    ("Annotation", "Annotation"),
]

ALL_HEADERS = TRACKING_COLUMNS + [col[1] for col in RVTOOLS_TO_OUTPUT]

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")


def sanitize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\ characters."""
    sanitized = re.sub(r'[\[\]:*?/\\]', '_', name)
    return sanitized[:31]


def extract_site_from_filename(filepath: str) -> str:
    """Extract site name from RVTools filename pattern like
    RVTools_export_all_2026-03-11_10.15.48_Site Beta.xlsx -> Site Beta
    """
    stem = Path(filepath).stem
    parts = stem.split("_")
    # Typical pattern: RVTools_export_all_YYYY-MM-DD_HH.MM.SS_SiteName[_extra]
    # Find the part after the time (HH.MM.SS pattern)
    for i, part in enumerate(parts):
        if re.match(r'^\d{2}\.\d{2}\.\d{2}$', part):
            site_parts = parts[i + 1:]
            if site_parts:
                return " ".join(site_parts)
    return stem


def read_rvtools_vinfo(filepath: str, exclude_powered_off: bool = False) -> list[dict]:
    """Read VMs from the vInfo sheet of an RVTools export."""
    wb = load_workbook(filepath, read_only=True, data_only=True)

    if "vInfo" not in wb.sheetnames:
        print(f"  WARNING: No 'vInfo' sheet in {filepath}, skipping.", file=sys.stderr)
        wb.close()
        return []

    ws = wb["vInfo"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    vms = []

    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        if not record.get("VM"):
            continue
        if exclude_powered_off and str(record.get("Powerstate", "")).lower() != "poweredon":
            continue
        vms.append(record)

    return vms


def determine_site(vm: dict, fallback_site: str) -> str:
    """Determine site from VM record, falling back to filename-derived site."""
    site = vm.get("Site Name")
    if site and str(site).strip():
        return str(site).strip()
    return fallback_site


def build_output_row(vm: dict) -> list:
    """Map a VM record to the output row (tracking cols blank + data cols)."""
    row = [""] * len(TRACKING_COLUMNS)

    for rvtools_col, output_col in RVTOOLS_TO_OUTPUT:
        if output_col == "Memory_GB":
            mem_mb = vm.get("Memory")
            try:
                row.append(round(float(mem_mb) / 1024, 2) if mem_mb else "")
            except (ValueError, TypeError):
                row.append("")
        elif output_col == "Disk_Capacity_GB":
            disk_mib = vm.get("Total disk capacity MiB")
            try:
                row.append(round(float(disk_mib) / 1024, 2) if disk_mib else "")
            except (ValueError, TypeError):
                row.append("")
        else:
            val = vm.get(rvtools_col, "")
            row.append(val if val is not None else "")

    return row


def write_schedule(site_vms: dict[str, list[dict]], output_path: str):
    """Write the migration schedule workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    for site_name in sorted(site_vms.keys()):
        vms = site_vms[site_name]
        sheet_name = sanitize_sheet_name(site_name)
        ws = wb.create_sheet(title=sheet_name)

        # Write header row
        for col_idx, header in enumerate(ALL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Write VM data rows
        for row_idx, vm in enumerate(vms, start=2):
            output_row = build_output_row(vm)
            for col_idx, val in enumerate(output_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-size columns (approximate)
        for col_idx in range(1, len(ALL_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            header_len = len(ALL_HEADERS[col_idx - 1])
            ws.column_dimensions[col_letter].width = max(header_len + 4, 12)

        # Freeze panes below header
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return len(site_vms)


def main():
    parser = argparse.ArgumentParser(
        description="Generate {CLIENT} Migration Weekly Schedule from RVTools exports"
    )
    parser.add_argument(
        "rvtools_files",
        nargs="+",
        help="One or more RVTools .xlsx export files",
    )
    parser.add_argument(
        "-o", "--output",
        default="Migration_Weekly_Schedule.xlsx",
        help="Output XLSX path (default: Migration_Weekly_Schedule.xlsx)",
    )
    parser.add_argument(
        "--exclude-powered-off",
        action="store_true",
        help="Exclude VMs that are not in poweredOn state",
    )
    args = parser.parse_args()

    site_vms: dict[str, list[dict]] = {}
    total_vms = 0

    for filepath in args.rvtools_files:
        if not Path(filepath).exists():
            print(f"WARNING: File not found: {filepath}", file=sys.stderr)
            continue
        if filepath.endswith(".zip"):
            print(f"  Skipping zip file: {filepath}", file=sys.stderr)
            continue

        fallback_site = extract_site_from_filename(filepath)
        print(f"Processing: {Path(filepath).name} (fallback site: {fallback_site})")

        vms = read_rvtools_vinfo(filepath, exclude_powered_off=args.exclude_powered_off)
        print(f"  Found {len(vms)} VMs")

        for vm in vms:
            site = determine_site(vm, fallback_site)
            site_vms.setdefault(site, []).append(vm)
            total_vms += 1

    if not site_vms:
        print("ERROR: No VMs found in any input file.", file=sys.stderr)
        sys.exit(1)

    num_sheets = write_schedule(site_vms, args.output)
    print(f"\nOutput: {args.output}")
    print(f"  {num_sheets} site sheets, {total_vms} total VMs")


if __name__ == "__main__":
    main()
